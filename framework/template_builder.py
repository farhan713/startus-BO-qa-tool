"""Build the official Stratus QA Test-Case template (.xlsx + .csv).

Produces two artifacts:
  • docs/Stratus-QA-TestCase-Template.xlsx — 3 sheets: Test Cases (empty,
    with data validation), Reference, Example.
  • docs/Stratus-QA-TestCase-Template.csv  — same column set, importable
    into Google Sheets via File → Import → Replace.

A QA fills the template, drops it in the tool, and the importer's fast
path produces 100% YAML with zero `todo` markers.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# The single source of truth for the template's column order + headers.
# The importer's fast-path also reads this list to detect the format.
COLUMNS = [
    ("TC ID",     16, "Stable ID — survives reorders. e.g. TC-CUST-001"),
    ("Scenario",  36, "Group label — repeated on every step row of the same test."),
    ("Screen",    24, "MUST match a catalog screenname (e.g. customerlist)."),
    ("Step",       8, "Per-scenario step order (1, 2, 3, …)."),
    ("Action",    18, "Pick from the dropdown."),
    ("Target",    34, "Element id (#Search), label (Last Name), or button text (Save)."),
    ("Value",     22, "For fill/select; the value typed or chosen."),
    ("Expected",  40, "Free text. Saved as a comment in YAML; not asserted."),
    ("Priority",  10, "P1 / P2 / P3 — optional."),
    ("Notes",     30, "Free text — saved as a YAML comment."),
]

ACTIONS = [
    "click", "fill", "select", "open_search", "wait",
    "assert_visible", "assert_text", "assert_no_errors",
    "assert_rows_min", "assert_rows_max", "screenshot",
]

PRIORITIES = ["P1", "P2", "P3"]

# Real example translated from STRAT_28795_TestCases.xlsx
EXAMPLE_ROWS = [
    # (TC ID, Scenario, Screen, Step, Action, Target, Value, Expected, Priority, Notes)
    ("TC-28795-01", "Shipping with correct size values", "shippingorderslist",
     1, "click",           "Ship Customer Orders", "",   "List opens", "P1", ""),
    ("TC-28795-01", "Shipping with correct size values", "shippingorderslist",
     2, "fill",            "Search",               "WEB-12345", "",  "P1", ""),
    ("TC-28795-01", "Shipping with correct size values", "shippingorderslist",
     3, "click",           "Search",               "",   "",            "P1", ""),
    ("TC-28795-01", "Shipping with correct size values", "shippingorderslist",
     4, "click",           "Do Shipping",          "",   "",            "P1", ""),
    ("TC-28795-01", "Shipping with correct size values", "shippingorderslist",
     5, "fill",            "To Ship",              "1",  "",            "P1", ""),
    ("TC-28795-01", "Shipping with correct size values", "shippingorderslist",
     6, "fill",            "Weight",               "1",  "",            "P1", ""),
    ("TC-28795-01", "Shipping with correct size values", "shippingorderslist",
     7, "fill",            "Length",               "5",  "",            "P1", ""),
    ("TC-28795-01", "Shipping with correct size values", "shippingorderslist",
     8, "fill",            "Width",                "5",  "",            "P1", ""),
    ("TC-28795-01", "Shipping with correct size values", "shippingorderslist",
     9, "fill",            "Height",               "5",  "",            "P1", ""),
    ("TC-28795-01", "Shipping with correct size values", "shippingorderslist",
     10, "click",          "Finish Packing",       "",   "",            "P1", ""),
    ("TC-28795-01", "Shipping with correct size values", "shippingorderslist",
     11, "assert_no_errors","",                    "",   "No error appears", "P1", ""),

    # Second test — incorrect values
    ("TC-28795-02", "Shipping with invalid size values", "shippingorderslist",
     1, "click",           "Ship Customer Orders", "",   "", "P1", ""),
    ("TC-28795-02", "Shipping with invalid size values", "shippingorderslist",
     2, "fill",            "Search",               "WEB-12345", "", "P1", ""),
    ("TC-28795-02", "Shipping with invalid size values", "shippingorderslist",
     3, "click",           "Search",               "",   "", "P1", ""),
    ("TC-28795-02", "Shipping with invalid size values", "shippingorderslist",
     4, "click",           "Do Shipping",          "",   "", "P1", ""),
    ("TC-28795-02", "Shipping with invalid size values", "shippingorderslist",
     5, "fill",            "Length",               "100", "", "P1", ""),
    ("TC-28795-02", "Shipping with invalid size values", "shippingorderslist",
     6, "click",           "Finish Packing",       "",   "", "P1", ""),
    ("TC-28795-02", "Shipping with invalid size values", "shippingorderslist",
     7, "assert_text",     "",                     "Requested Dimensions",
     "Error message visible", "P1", "Validates fix for STRAT-28795"),

    # Customer-list smoke
    ("TC-CUST-001", "Customer search by Last Name",     "customerlist",
     1, "click",           "Customer Management",  "",   "Menu hover", "P2", ""),
    ("TC-CUST-001", "Customer search by Last Name",     "customerlist",
     2, "click",           "Customer List",        "",   "", "P2", ""),
    ("TC-CUST-001", "Customer search by Last Name",     "customerlist",
     3, "open_search",     "",                     "",   "", "P2", ""),
    ("TC-CUST-001", "Customer search by Last Name",     "customerlist",
     4, "fill",            "Last Name",            "Smith", "", "P2", ""),
    ("TC-CUST-001", "Customer search by Last Name",     "customerlist",
     5, "click",           "Search",               "",   "", "P2", ""),
    ("TC-CUST-001", "Customer search by Last Name",     "customerlist",
     6, "assert_no_errors","",                     "",   "", "P2", ""),
]

# ---------- styling helpers ----------
_HEADER_FILL = PatternFill("solid", start_color="0F62FE")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_REFHEAD_FILL = PatternFill("solid", start_color="161616")
_THIN = Side(style="thin", color="C6C6C6")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _write_test_cases_sheet(ws, rows: list | None = None) -> None:
    """The blank Test Cases sheet (with header + data validation)."""
    ws.title = "Test Cases"
    ws.append([c[0] for c in COLUMNS])
    for i, (name, width, _) in enumerate(COLUMNS, 1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = width
        cell = ws.cell(row=1, column=i)
        cell.fill = _HEADER_FONT and _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = _BORDER

    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    # Action dropdown — column E (5)
    dv_action = DataValidation(type="list",
                               formula1='"' + ",".join(ACTIONS) + '"',
                               allow_blank=True,
                               error="Pick a valid action.",
                               errorTitle="Invalid action")
    ws.add_data_validation(dv_action)
    dv_action.add(f"E2:E10000")

    # Priority dropdown — column I (9)
    dv_pri = DataValidation(type="list",
                            formula1='"' + ",".join(PRIORITIES) + '"',
                            allow_blank=True)
    ws.add_data_validation(dv_pri)
    dv_pri.add(f"I2:I10000")

    # Step # numeric only — column D (4)
    dv_step = DataValidation(type="whole", operator="greaterThan", formula1="0",
                             allow_blank=True, error="Step must be a positive integer.")
    ws.add_data_validation(dv_step)
    dv_step.add(f"D2:D10000")

    # Pre-fill example rows if requested
    if rows:
        for r in rows:
            ws.append(r)
        # apply borders + word wrap to data cells
        for r in range(2, ws.max_row + 1):
            for c in range(1, len(COLUMNS) + 1):
                ws.cell(row=r, column=c).border = _BORDER
                ws.cell(row=r, column=c).alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True)


def _write_reference_sheet(ws) -> None:
    ws.title = "Reference"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30

    title = ws.cell(row=1, column=1, value="Allowed Actions — quick reference")
    title.font = Font(name="Calibri", bold=True, size=14, color="161616")
    ws.merge_cells("A1:D1")

    headers = ["Action", "When to use", "Target", "Value"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.fill = _REFHEAD_FILL
        c.font = Font(name="Calibri", bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="left")

    rows = [
        ("click",            "Click a button or link",         '#buttonId · "Save" · New',  "(blank)"),
        ("fill",             "Enter text into a field",        '#fieldId · "Last Name"',    "the value to type"),
        ("select",           "Pick from a dropdown",           '#fieldId · label',          "option label"),
        ("open_search",      "Reveal the search criteria pane","(blank)",                   "(blank)"),
        ("wait",             "Pause N seconds",                "N (e.g. 2)",                "(blank)"),
        ("assert_visible",   "Element must be visible",        '#element · "text"',         "(blank)"),
        ("assert_text",      "Page body must contain text",    "(blank)",                   "exact text"),
        ("assert_no_errors", "No error words on the page",     "(blank)",                   "(blank)"),
        ("assert_rows_min",  "Grid has at least N rows",       "(blank)",                   "N"),
        ("assert_rows_max",  "Grid has at most N rows",        "(blank)",                   "N"),
        ("screenshot",       "Capture a screenshot",           "filename suffix",           "(blank)"),
    ]
    for r_i, row in enumerate(rows, start=4):
        for c_i, val in enumerate(row, 1):
            c = ws.cell(row=r_i, column=c_i, value=val)
            c.border = _BORDER
            if c_i == 1:
                c.font = Font(name="Calibri", bold=True)
                c.fill = PatternFill("solid", start_color="EDF5FF")

    # Authoring tips
    tip_row = len(rows) + 6
    ws.cell(row=tip_row, column=1, value="Tips").font = Font(bold=True, size=12)
    tips = [
        "• Use ONE ROW PER STEP, not one row per scenario. Repeat TC ID + Scenario on every step.",
        "• Screen MUST match a catalog screenname (e.g. customerlist, shippingorderslist).",
        '• Target accepts #id, label text in quotes, or button text. e.g.  #Search   "Last Name"   Save',
        "• Step numbers within a Scenario decide run order. Order across TC IDs doesn't matter.",
        "• Empty Target/Value is fine — the action doesn't always need both.",
        "• Expected is informational only; use assert_* actions if you want it checked.",
    ]
    for i, tip in enumerate(tips):
        c = ws.cell(row=tip_row + 1 + i, column=1, value=tip)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.merge_cells(start_row=tip_row + 1 + i, start_column=1,
                       end_row=tip_row + 1 + i,   end_column=4)


def build_template(out_dir: Path) -> tuple[Path, Path]:
    """Write the .xlsx template + a CSV copy for Google Sheets import."""
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    _write_test_cases_sheet(wb.active)                  # Sheet 1 — blank template
    _write_reference_sheet(wb.create_sheet("Reference"))  # Sheet 2 — cheat sheet
    _write_test_cases_sheet(wb.create_sheet("Example"), EXAMPLE_ROWS)  # Sheet 3 — filled

    xlsx_path = out_dir / "Stratus-QA-TestCase-Template.xlsx"
    wb.save(xlsx_path)

    # CSV: column headers + one example row so Google Sheets users get the
    # structure on Import → Replace.
    csv_path = out_dir / "Stratus-QA-TestCase-Template.csv"
    with csv_path.open("w", encoding="utf-8") as f:
        f.write(",".join(c[0] for c in COLUMNS) + "\n")
        for r in EXAMPLE_ROWS[:3]:                       # 3 sample rows
            cells = []
            for v in r:
                s = str(v).replace('"', '""')
                cells.append(f'"{s}"' if any(c in s for c in ',"\n') else s)
            f.write(",".join(cells) + "\n")

    return xlsx_path, csv_path


if __name__ == "__main__":
    xlsx, csv = build_template(Path(__file__).resolve().parents[1] / "docs")
    print(f"Wrote {xlsx}")
    print(f"Wrote {csv}")
