"""Excel test-case importer.

Converts the kind of Excel test-case files Celerant QA teams produce into the
YAML format `parse_custom_tests` already accepts. Three real-world layouts
are supported (auto-detected from header text), and prose steps are
translated to executable actions by a deterministic rule engine.

Layouts handled (header keywords used to detect):
  A. FLAT       — one row per test: TEST CASE | FEATURE | TEST STEPS | EXPECTED ...
  B. GROUPED    — one row per scenario "TSn: ..." then numbered step rows below
                  with blank scenario column (the Test_Case_27857 style).
  C. NUMBERED   — TC No | Test Case Scenario | Steps | Expected ... (the
                  customerNotification style with 76 numbered tests).

Anything the translator can't map confidently becomes a `{action: todo, ...}`
step so the YAML round-trips cleanly; the runner logs it as a TODO line.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field


# ============================================================ Translator
# All rules are deterministic regexes. Order matters — first match wins.
# Capture groups feed the action's target/value.

#: A friendly button name (English noun phrase) — used to extract click
#: targets from prose like "Click Save", "Hit OK", "Press Search".
_BTN_PHRASE = r"[A-Z][\w &/.\-]{1,40}?"

#: A field name — anything that looks like an identifier or quoted phrase
_FIELD = r"['\"]?([A-Za-z][\w \-./]{1,40})['\"]?"

#: A value — quoted, equals-quoted, or after "with"
_VALUE = r"['\"]([^'\"]{1,60})['\"]|=\s*([\w\-./]{1,40})"


# Each rule: (regex, action-builder). Builder returns dict (or None to skip).
def _r_navigate(m):
    target = m.group(1).strip().rstrip(".").rstrip(",")
    # Stratus breadcrumb shapes:  A > B > C   /   A->B->C   /   A>>B>>C
    target = re.sub(r"\s*(?:>>|->|>|→)\s*", " > ", target)
    if " > " in target:
        target = target.split(" > ")[-1].strip()
    return {"action": "click", "target": target, "_hint": "navigate"}

def _r_click(m):
    return {"action": "click", "target": m.group(1).strip().rstrip(".")}

def _r_fill(m):
    val = m.group(1).strip().strip("'\"")
    field = m.group(2).strip().strip("'\"")
    return {"action": "fill", "target": field, "value": val}

def _r_fill_rev(m):
    field = m.group(1).strip().strip("'\"")
    val = m.group(2).strip().strip("'\"")
    return {"action": "fill", "target": field, "value": val}

def _r_select(m):
    val = m.group(1).strip().strip("'\"")
    field = m.group(2).strip().strip("'\"")
    return {"action": "select", "target": field, "value": val}

def _r_check(m):
    return {"action": "click", "target": m.group(1).strip(), "_hint": "checkbox"}

def _r_save(_):  return {"action": "click", "target": "Save"}
def _r_cancel(_): return {"action": "click", "target": "Cancel"}
def _r_search(_): return {"action": "click", "target": "Search"}
def _r_reset(_):  return {"action": "click", "target": "Reset"}
def _r_new(_):    return {"action": "click", "target": "New"}
def _r_edit(_):   return {"action": "click", "target": "Edit"}
def _r_delete(_): return {"action": "click", "target": "Delete"}
def _r_print(_):  return {"action": "click", "target": "Print"}

def _r_assert_visible(m):
    return {"action": "assert_visible", "target": m.group(1).strip().rstrip(".")}
def _r_assert_not_visible(m):
    # We have no `assert_not_visible` action yet — keep as todo so a human
    # can decide between checking absence vs simply not asserting presence.
    return {"action": "todo",
            "target": f"verify NOT visible: {m.group(1).strip()}",
            "_hint": "assert_not_visible"}
def _r_assert_text(m):
    return {"action": "assert_text", "value": m.group(1).strip().rstrip(".")}
def _r_assert_no_errors(_):
    return {"action": "assert_no_errors"}

def _r_login(_):
    return {"action": "todo", "target": "ensure logged in (handled by tool's authentication step)"}


# Common step-line prefix: optional enumeration ("1.", "1)", "2:", "-", "•")
_PFX = r"\s*(?:[\-\*•]|\d+[\.\)\:])?\s*"

# Captures a Stratus breadcrumb tail: "Back Office > Admin > Set Enterprise"
# (matches  >  ,  ->  ,  >>  ,  →  ).
_PATH_TAIL = r"(?:Back\s*Office\s*(?:>>|->|>|→)\s*)?(.+?)"


# (regex, builder, label) — first match wins. Tested case-insensitively.

# ---- Legacy arrow-chain segment builders (STRAT test sheets) -------------
# A split "Path: ... -> search -> select -> Put Tag type -> save" leaves bare
# segments that carry no verb. They are still perfectly typed steps: a known
# button name is a click, "X Tab" is a click, and "Put <field>" is a fill whose
# value the tester supplies. Emitting a typed step with an empty value beats a
# todo, because the editor can then show a value box instead of raw prose.
_BARE_BUTTONS = ("search","save","edit","delete","print","new","reset","close",
                 "cancel","select","action","add","update","copy","submit",
                 "apply","ok","next","back","refresh","export","import")

def _r_bare_button(m):
    return {"action": "click", "target": m.group(1).strip().title()}

def _r_tab(m):
    return {"action": "click", "target": m.group(1).strip()}

def _r_put_field(m):
    field = m.group(1).strip().rstrip(".").rstrip(",")
    # Trailing parenthetical is guidance for the tester, not part of the name:
    # "New Tag (Yes/No)" -> field "New Tag".
    field = re.sub(r"\s*\([^)]*\)\s*$", "", field).strip()
    return {"action": "fill", "target": field, "value": ""}


def _r_named_button(m):
    """"Search - Searches using selected options." -> click Search.

    Legacy sheets document a screen's buttons as "<Button> - <what it does>".
    The name before the dash is a real control, so this is a click, not prose.
    Guarded to short names so "Buy/Trade Deal OH Aggregate 1 - 5" (a field whose
    label happens to contain a dash) is not mistaken for a button."""
    return {"action": "click", "target": m.group(1).strip()}

def _r_verify_state(_):
    return {"action": "assert_no_errors", "target": ""}

RULES: list = [
    # ---- DB/state assertions — keep as todo with the prose preserved ------
    (re.compile(r"\btb_\w+\b", re.I),
     lambda m: {"action": "todo", "target": f"DB check: {m.string.strip()[:160]}",
                "_hint": "db_state"}, "db-state"),
    (re.compile(r"^" + _PFX + r"(?:Make sure|Confirm|Verify|Ensure)\b.*", re.I),
     lambda m: {"action": "todo", "target": f"verify state: {m.string.strip()[:160]}",
                "_hint": "verify"}, "verify-prose"),

    # ---- Login / logout ---------------------------------------------------
    (re.compile(r"^" + _PFX + r"(?:Log ?in|Sign in)\b.*$", re.I), _r_login, "login"),

    # ---- Navigation — breadcrumb path-based first -------------------------
    # "1. vendor management -> purchase order"
    # "Back Office > Admin > Set Enterprise > Sales / POS Tab"
    # "Sale Tracking -> Receipts -> search"
    (re.compile(r"^" + _PFX + r"(?:Go to|Navigate to|Open|From|In|Login to|Go back to)?\s*"
                r"(?:Back\s*Office\s*(?:>>|->|>|→)\s*)?"
                r"([\w &/.\-]+(?:\s*(?:>>|->|>|→)\s*[\w &/.\-]+)+)\s*\.?\s*$", re.I),
     _r_navigate, "navigate-path"),
    # "Go to BackOffice menu" / "Open Customer screen"
    (re.compile(r"^" + _PFX + r"(?:Go to|Navigate to|Open)\s+(?:the\s+)?(.+?)"
                r"(?:\s+(?:menu|tab|screen|page|module|section))?\s*\.?\s*$", re.I),
     _r_navigate, "navigate-named"),

    # ---- Save / Cancel / Search etc. (idiomatic buttons) ------------------
    (re.compile(r"^" + _PFX + r"(?:Click|Hit|Press|Tap)\s+(?:on\s+)?(?:the\s+)?Save\b.*$", re.I),   _r_save,   "save"),
    (re.compile(r"\b(?:and\s+)?save\s+it\b\.?\s*$", re.I),                                          _r_save,   "save-tail"),
    (re.compile(r"^" + _PFX + r"(?:Click|Hit|Press)\s+(?:on\s+)?(?:the\s+)?Cancel\b.*$", re.I),     _r_cancel, "cancel"),
    (re.compile(r"^" + _PFX + r"(?:Click|Hit|Press)\s+(?:on\s+)?(?:the\s+)?Search\b.*$", re.I),     _r_search, "search"),
    (re.compile(r"^" + _PFX + r"(?:Click|Hit|Press)\s+(?:on\s+)?(?:the\s+)?Reset\b.*$", re.I),      _r_reset,  "reset"),
    (re.compile(r"^" + _PFX + r"(?:Click|Hit|Press)\s+(?:on\s+)?(?:the\s+)?(?:New|Add)\b.*$", re.I), _r_new,   "new"),
    (re.compile(r"^" + _PFX + r"(?:Click|Hit|Press)\s+(?:on\s+)?(?:the\s+)?Edit\b.*$", re.I),       _r_edit,   "edit"),
    (re.compile(r"^" + _PFX + r"(?:Click|Hit|Press)\s+(?:on\s+)?(?:the\s+)?Delete\b.*$", re.I),     _r_delete, "delete"),
    (re.compile(r"^" + _PFX + r"(?:Click|Hit|Press)\s+(?:on\s+)?(?:the\s+)?Print\b.*$", re.I),      _r_print,  "print"),
    # "Hit OK" / "Click OK button"
    (re.compile(r"^" + _PFX + r"(?:Click|Hit|Press|Tap)\s+(?:on\s+)?(?:the\s+)?(OK|Yes|No|Continue|Apply|Submit)\b.*$", re.I),
     _r_click, "hit-confirm"),

    # ---- Generic click on a named element ---------------------------------
    (re.compile(r"^" + _PFX + r"(?:Click|Hit|Press|Tap|Choose|Select)\s+(?:on\s+)?(?:the\s+)?"
                r"(" + _BTN_PHRASE + r")(?:\s+(?:button|link|tab|option))?\s*\.?\s*$", re.I),
     _r_click, "click-named"),

    # ---- Check / uncheck checkboxes --------------------------------------
    (re.compile(r"^" + _PFX + r"(?:Check|Uncheck|Tick|Untick)\s+(?:on\s+)?(?:the\s+)?"
                r"(.+?)(?:\s+(?:checkbox|option|box|flag))?\s*\.?\s*$", re.I),
     _r_check, "check"),

    # ---- Fill / type / enter ---------------------------------------------
    # "Enter 'QA' in Last Name"   /   "Type 'QA' into the search field"
    (re.compile(r"^" + _PFX + r"(?:Enter|Type|Input|Fill)\s+(?:in\s+)?"
                r"['\"]([^'\"]{1,60})['\"]\s+(?:in|into|to|for)\s+(?:the\s+)?(.+?)"
                r"(?:\s+(?:field|input|box|column))?\s*\.?\s*$", re.I),
     _r_fill, "fill"),
    # "Fill in Last Name with 'QA'"  /  "Set Status to Active"
    (re.compile(r"^" + _PFX + r"(?:Fill in|Set)\s+(?:the\s+)?(.+?)\s+(?:with|to|=)\s+"
                r"['\"]?([^'\"]{1,60})['\"]?\s*\.?\s*$", re.I),
     _r_fill_rev, "fill-rev"),

    # ---- Shorthand:  Field = Value   /   Field: Value -------------------
    # Celerant tests often list field assignments as bare "To Ship=1",
    # "Weight=1.5", "Length=100", etc. Match conservatively: short LHS that
    # looks like a field label, short RHS that looks like a literal value.
    (re.compile(r"^" + _PFX + r"([A-Z][A-Za-z][\w \-]{0,30})\s*[=:]\s*([\w./\-]{1,40})\s*\.?\s*$"),
     _r_fill_rev, "fill-shorthand"),

    # ---- Select dropdowns -------------------------------------------------
    # "Select 'Active' from Status dropdown"
    (re.compile(r"^" + _PFX + r"(?:Select|Choose|Pick)\s+"
                r"['\"]([^'\"]{1,60})['\"]\s+(?:from|in)\s+(?:the\s+)?(.+?)"
                r"(?:\s+(?:dropdown|select|menu|list|option))?\s*\.?\s*$", re.I),
     _r_select, "select"),
    # "Select Active from Status dropdown" (no quotes — only if 'dropdown' present)
    (re.compile(r"^" + _PFX + r"(?:Select|Choose|Pick)\s+(.+?)\s+(?:from|in)\s+(?:the\s+)?(.+?)"
                r"\s+(?:dropdown|select|list)\s*\.?\s*$", re.I),
     _r_select, "select-noquote"),

    # ---- Assertions (negative first so 'should not' beats 'should') ------
    (re.compile(r"\b(?:should not|shouldn['’]?t|must not|cannot)\s+(?:be\s+)?(?:visible|shown|displayed|appear|allowed|available)\b.*$", re.I),
     _r_assert_not_visible, "assert-not-visible"),
    (re.compile(r"\b(?:saves?\s+(?:properly|successfully|correctly)|saved?\s+correctly|saves? to (?:the )?(?:database|db))\b.*$", re.I),
     _r_assert_no_errors, "assert-no-errors"),
    (re.compile(r"\bno (?:errors?|warning)s?\b.*$", re.I),
     _r_assert_no_errors, "assert-no-errors"),
    # "should be visible" / "should display X" / "should appear"
    (re.compile(r"\b(?:should|must|will)\s+(?:be\s+)?(?:visible|shown|displayed|appear|present|available)\b.*$", re.I),
     _r_assert_visible, "assert-visible"),

    # ---- Display / show a literal quoted text -----------------------------
    (re.compile(r"\b(?:displays?|shows?|reads?|says?)\s+['\"]([^'\"]{2,80})['\"]", re.I),
     _r_assert_text, "assert-text"),

    # ---- Legacy sheet shapes -------------------------------------------
    # "verify state: Make sure all the details are saved properly"
    (re.compile(r"^\s*verify\s+state\s*:", re.I), _r_verify_state, "verify-state"),
    # "<Button> - <description>"  (<=3 words before the dash)
    (re.compile(r"^\s*([A-Z][\w/]*(?:\s+[A-Z][\w/]*){0,2})\s+-\s+[A-Z].{4,}$"),
     _r_named_button, "named-button"),

    # ---- Bare arrow-chain segments (must stay last: lowest precedence) ----
    (re.compile(r"^\s*(" + "|".join(_BARE_BUTTONS) + r")\s*\.?\s*$", re.I),
     _r_bare_button, "bare-button"),
    (re.compile(r"^\s*(.{2,40}?)\s+tabs?\s*\.?\s*$", re.I),
     _r_tab, "bare-tab"),
    (re.compile(r"^\s*(?:Put|Enter|Set|Type|Input|Give|Provide)\s+(?:the\s+)?(?:value\s+of\s+)?(.{2,48}?)\s*\.?\s*$", re.I),
     _r_put_field, "put-field"),
]


def translate_step(prose: str) -> dict:
    """Translate one English step into an action dict, or a `todo` marker.

    Always returns a dict — even unmappable prose becomes
    `{action: todo, target: <original>}` so it survives round-trip and is
    visible in the run log."""
    p = (prose or "").strip()
    if not p:
        return {"action": "todo", "target": "(empty step)"}
    # Strip leading enumeration like "1." / "1)" / "1:" / "-"
    p_norm = re.sub(r"^\s*(?:[\-\*•]|\d+[\.\)\:])\s*", "", p)
    # "Path: A > B > C" is the sheets' way of writing navigation; the bare
    # breadcrumb is what navigate-path already understands.
    p_norm = re.sub(r"^\s*path\s*:\s*", "", p_norm, flags=re.I)
    for rx, builder, label in RULES:
        m = rx.search(p_norm) if label == "db-state" else rx.match(p_norm)
        if m:
            try:
                out = builder(m)
                if out:
                    # Drop hint key — it's only for debugging
                    out.pop("_hint", None)
                    return out
            except Exception:
                pass
    # Unmappable — preserve the original prose verbatim
    return {"action": "todo", "target": p_norm[:200]}


# ============================================================ Layout parsing

@dataclass
class TestCase:
    name: str
    steps: list                          # list of action dicts
    expected: str = ""
    notes: str = ""                      # raw prose for the YAML comment
    lines: list = field(default_factory=list)      # the prose lines, in order


def _norm(s) -> str:
    if s is None: return ""
    return str(s).replace("\xa0", " ").strip()


_ARROW_SPLIT = re.compile(r"\s*->\s*")
_PATH_HEAD   = re.compile(r"^\s*path\s*:\s*(.+)$", re.I)


def _split_arrow_chain(chunk: str) -> list[str]:
    """Expand a "Path: A > B > C -> do X -> do Y" cell into separate steps.

    Legacy Celerant test sheets write a whole test on one line: '>' walks the
    menu, then each '->' is the next thing the tester does. Left whole it
    matches no rule and lands as a single unusable todo, which is most of what
    the 8% translation rate on these files actually was. Splitting first lets
    every segment hit the normal rules; the breadcrumb head keeps its '>' so
    _r_navigate still resolves it to the destination screen."""
    if "->" not in chunk:
        return [chunk]
    parts = [p.strip() for p in _ARROW_SPLIT.split(chunk) if p.strip()]
    if len(parts) < 2:
        return [chunk]
    head, rest = parts[0], parts[1:]
    m = _PATH_HEAD.match(head)
    if m:
        head = "Go to " + m.group(1).strip()
    return [head] + rest


def _split_steps(text: str) -> list[str]:
    """Split a Steps cell into individual numbered/bulleted lines."""
    return [s for s, _ in _split_steps_src(text)]


def _split_steps_src(text: str) -> list[tuple]:
    """Same split, but each step carries the index of the prose line it came from.

    Provenance is what lets the editor group steps under the sentence the tester
    actually wrote, and — more importantly — show a line that produced NO steps.
    A silently dropped requirement is invisible in any layout that only lists
    what was produced.
    """
    if not text: return []
    lines = _source_lines(text)
    out = []
    for idx, line in enumerate(lines):
        for piece in _split_arrow_chain(line):
            out.append((piece, idx))
    return out


def _source_lines(text: str) -> list[str]:
    """The prose lines of a cell, as the tester wrote them."""
    if not text: return []
    raw = re.split(r"(?:\r?\n)+|(?<=[\.\)])\s+(?=\d+[\.\)])", str(text))
    return [c.strip() for c in raw if c.strip()]


def _detect_layout(rows: list[list[str]]) -> tuple[str, int]:
    """Find the header row + return ('flat'|'grouped'|'numbered', header_row_index)."""
    for i, row in enumerate(rows[:15]):
        joined = " | ".join(_norm(c).lower() for c in row)
        if "test step" in joined or "steps" in joined:
            if "test scenario" in joined or "tc no" in joined or "test case" in joined:
                if "tc no" in joined: return "numbered", i
                return "flat", i
            return "flat", i
    return "flat", 0


def _index_columns(header_row: list[str]) -> dict:
    """Map column index to a normalized name we care about."""
    out = {}
    for i, cell in enumerate(header_row):
        n = _norm(cell).lower()
        if not n: continue
        if "tc no" in n or n == "test case":
            out["num"] = i
        elif "scenario" in n or "feature" in n:
            out["scenario"] = i
        elif "step" in n:
            out["steps"] = i
        elif "expected" in n or "expectation" in n:
            out["expected"] = i
        elif "comment" in n or "summary" in n:
            out["comments"] = i
        elif "priority" in n:
            out["priority"] = i
    return out


def _parse_flat(rows, cols, header_idx) -> list[TestCase]:
    """Rows grouped under their Test Scenario.

    A blank scenario cell means "still the scenario above" — these sheets write
    one scenario and then continue its steps down the following rows. Treating
    every row as its own test shattered the file: the Services sheet has 34 test
    cases across 221 rows, but row-per-test turned that into 171 "scenarios",
    most holding a single step, with headings like "Buttons" promoted to test
    names. Carrying the last name forward keeps a scenario whole and puts its
    steps inside it, which is what the sheet actually says.
    """
    cases: list[TestCase] = []
    for r in rows[header_idx + 1:]:
        scenario = _norm(r[cols["scenario"]]) if "scenario" in cols and cols["scenario"] < len(r) else ""
        steps_text = _norm(r[cols["steps"]])  if "steps"    in cols and cols["steps"]    < len(r) else ""
        expected   = _norm(r[cols["expected"]]) if "expected" in cols and cols["expected"] < len(r) else ""
        if not (scenario or steps_text or expected): continue
        # Skip preamble rows like "Author: X" or "Prerequisite: ..."
        if scenario.lower().startswith(("author:", "prerequisite", "prerequesite", "note:")):
            continue

        # Steps keep the index of the prose line they came from. Expected-result
        # lines continue the same numbering so both halves of the row are
        # addressable from one list.
        lines = _source_lines(steps_text)
        exp_lines = _source_lines(expected)
        steps = []
        for txt, si in _split_steps_src(steps_text):
            st = translate_step(txt); st["_src"] = si; st["_origin"] = "sheet"
            steps.append(st)
        for txt, si in _split_steps_src(expected):
            st = translate_step(txt); st["_src"] = len(lines) + si; st["_origin"] = "expected"
            steps.append(st)
        all_lines = lines + exp_lines

        # A sheet with a "TEST CASE"/"TC No" column numbers its cases, and that
        # number is the real boundary: the Services sheet has 34 numbered cases
        # but 44 rows carrying a FEATURE name, because one case spans several
        # feature lines. Where a number exists it wins; otherwise fall back to
        # "a new name starts a new case".
        num = _norm(r[cols["num"]]) if "num" in cols and cols["num"] < len(r) else ""
        starts_new = bool(num) if "num" in cols else bool(scenario)

        if starts_new or not cases:
            cases.append(TestCase(
                name=scenario or steps_text[:60] or f"test {num or len(cases)+1}",
                steps=steps,
                expected=expected,
                notes=steps_text,
                lines=all_lines,
            ))
        else:
            tc = cases[-1]
            # A continuation row may still name a sub-feature; keep it as a step
            # label so the detail is not lost when the rows are merged.
            base = len(tc.lines)
            if scenario and scenario not in (tc.name or ""):
                steps = [{"action": "todo", "target": scenario,
                          "_src": base, "_origin": "sheet"}] + steps
                all_lines = [scenario] + all_lines
                for s in steps[1:]:
                    s["_src"] = s.get("_src", 0) + 1
            for s in steps:
                s["_src"] = s.get("_src", 0) + base
            tc.lines.extend(all_lines)
            tc.steps.extend(steps)
            if expected:
                tc.expected = (tc.expected + "\n" + expected).strip() if tc.expected else expected
            if steps_text:
                tc.notes = (tc.notes + "\n" + steps_text).strip() if tc.notes else steps_text
    return cases


def _parse_grouped(rows, cols, header_idx) -> list[TestCase]:
    """Scenario header row, then 1+ step rows with blank scenario."""
    cases = []
    current = None
    for r in rows[header_idx + 1:]:
        scenario = _norm(r[cols["scenario"]]) if "scenario" in cols and cols["scenario"] < len(r) else ""
        steps_text = _norm(r[cols["steps"]])  if "steps"    in cols and cols["steps"]    < len(r) else ""
        expected   = _norm(r[cols["expected"]]) if "expected" in cols and cols["expected"] < len(r) else ""
        if scenario.lower().startswith(("author:", "prerequisite", "prerequesite", "note:")):
            continue
        if scenario:
            # New scenario starts
            if current: cases.append(current)
            current = TestCase(name=scenario, steps=[], expected="", notes="")
        if current is None:
            if not (steps_text or expected): continue
            current = TestCase(name=steps_text[:60] or f"test {len(cases)+1}", steps=[], expected="", notes="")
        for s in _split_steps(steps_text):
            current.steps.append(translate_step(s))
        for e in _split_steps(expected):
            current.steps.append(translate_step(e))
        if expected:
            current.expected = (current.expected + " | " + expected).strip(" |")
    if current: cases.append(current)
    return cases


def _looks_grouped(rows, cols, header_idx) -> bool:
    """Detect the GROUPED pattern: scenario column has 'TSn:' headers
    followed by blank scenario rows."""
    if "scenario" not in cols: return False
    has_ts_header = False
    has_blank_followup = False
    seen_ts = False
    for r in rows[header_idx + 1:header_idx + 20]:
        if cols["scenario"] >= len(r): continue
        s = _norm(r[cols["scenario"]])
        if re.match(r"^TS\d+:", s, re.I):
            has_ts_header = True
            seen_ts = True
        elif seen_ts and not s and "steps" in cols and cols["steps"] < len(r) and _norm(r[cols["steps"]]):
            has_blank_followup = True
    return has_ts_header and has_blank_followup


# ============================================================ Template fast-path

# Headers from `framework.template_builder.COLUMNS` (kept in sync).
TEMPLATE_HEADERS = ["TC ID", "Scenario", "Screen", "Step", "Action",
                    "Target", "Value", "Expected", "Priority", "Notes"]


def _looks_like_template(rows: list) -> int | None:
    """Return the header row index if the sheet is in the official template
    format, else None."""
    for i, row in enumerate(rows[:8]):
        cells = [_norm(c).lower() for c in row]
        if (cells[:5] == [h.lower() for h in TEMPLATE_HEADERS[:5]]):
            return i
        # Tolerant match: TC ID + Scenario + Screen + Step + Action all present
        # somewhere in the first 12 columns
        wanted = {h.lower() for h in TEMPLATE_HEADERS[:5]}
        if wanted.issubset(set(cells[:12])):
            return i
    return None


def _parse_template(rows: list, header_idx: int) -> dict:
    """Read the template format directly into a {screen: [TestCase,...]}
    mapping. No prose translation needed — each row IS a step."""
    header = [_norm(c).lower() for c in rows[header_idx]]
    # Tolerant column lookup
    def col(name: str) -> int | None:
        try: return header.index(name.lower())
        except ValueError: return None
    c_id, c_scn, c_scr, c_step, c_act = (
        col("TC ID"), col("Scenario"), col("Screen"), col("Step"), col("Action"))
    c_tgt, c_val, c_exp = col("Target"), col("Value"), col("Expected")

    # Group: {screen: {scenario_key: TestCase}}
    grouped: dict[str, dict[str, TestCase]] = {}
    # `scenario_key` is "TC ID || Scenario name" — TC ID takes precedence so
    # rows in any order group correctly.

    for r in rows[header_idx + 1:]:
        # Skip blank rows
        if not any(_norm(c) for c in r): continue
        screen = _norm(r[c_scr]) if c_scr is not None and c_scr < len(r) else ""
        action = _norm(r[c_act]) if c_act is not None and c_act < len(r) else ""
        if not screen or not action: continue          # row needs at least these
        tcid    = _norm(r[c_id])  if c_id  is not None and c_id  < len(r) else ""
        scenario = _norm(r[c_scn]) if c_scn is not None and c_scn < len(r) else ""
        step_n  = _norm(r[c_step]) if c_step is not None and c_step < len(r) else ""
        target  = _norm(r[c_tgt]) if c_tgt is not None and c_tgt < len(r) else ""
        value   = _norm(r[c_val]) if c_val is not None and c_val < len(r) else ""
        expected = _norm(r[c_exp]) if c_exp is not None and c_exp < len(r) else ""

        key = (tcid or "") + "||" + (scenario or "")
        sm = grouped.setdefault(screen, {})
        if key not in sm:
            sm[key] = TestCase(
                name=(scenario or tcid or f"test {len(sm)+1}")[:120],
                steps=[], expected=expected, notes=tcid)

        step: dict = {"action": action.lower()}
        if target: step["target"] = target
        if value:  step["value"] = value
        # Stamp the per-step number so we can sort, then drop it
        try: step["_n"] = int(float(step_n)) if step_n else len(sm[key].steps) + 1
        except Exception: step["_n"] = len(sm[key].steps) + 1
        sm[key].steps.append(step)

    # Sort each scenario's steps by _n then strip the helper key
    for sm in grouped.values():
        for tc in sm.values():
            tc.steps.sort(key=lambda s: s.get("_n", 0))
            for s in tc.steps:
                s.pop("_n", None)

    return grouped



# ============================================ inventory -> presence checks

_HEADING = re.compile(
    r"^\s*(?:buttons?|fields?|grid|grids?\s+results?|search\s+(?:criteria|components?)"
    r"|navigation\s+buttons?|columns?|bottom\s+(?:section|information)"
    r"|top\s+section|customer\s+info|deal\s+grid(?:\s+buttons?)?"
    r"|components?|tabs?|menu\s+options?)\s*:?\s*$", re.I)

_VERBY = re.compile(
    r"^\s*(?:click|select|choose|enter|put|type|set|open|go|search|save|add|check|"
    r"verify|scan|create|edit|delete|update|close|print|apply|enable|disable|"
    r"login|navigate|press|tap|fill|remove|view)\b", re.I)

_BREADCRUMBISH = re.compile(r"(?:>>|->|\s>\s|→)")


def _infer_inventory_checks(steps: list) -> int:
    """Turn a screen's component list into presence assertions.

    These sheets document a screen by listing what it contains:

        Search Components:      <- heading
          Deal Number           <- element
          Store
          Payout Method

    Those rows are not instructions, which is why they arrive as `todo` and get
    raised as questions the tester cannot meaningfully answer. But the list IS
    testable: every named element should be on screen. Converting each to
    assert_visible turns a wall of questions into real automated checks.

    Only rows directly under a heading are converted, and only short noun
    phrases — anything with a verb, a breadcrumb, or sentence length is left
    alone so genuine steps and prose are never rewritten.
    """
    changed, in_list = 0, False
    for st in steps:
        if (st.get("action") or "") != "todo":
            in_list = False
            continue
        text = (st.get("target") or "").strip()
        if not text:
            continue
        if _HEADING.match(text):
            in_list = True
            continue
        if not in_list:
            continue
        if (len(text) > 40 or _VERBY.match(text) or _BREADCRUMBISH.search(text)
                or text.endswith((".", ";", ":"))):
            in_list = False
            continue
        st["action"] = "assert_visible"
        st["target"] = text
        changed += 1
    return changed



_CONTROLish = re.compile(r"^[A-Z][A-Za-z0-9]*(?:[ /&\-][A-Za-z0-9()][A-Za-z0-9()]*){0,3}$")
_NOISE      = re.compile(r"[=:;\.]|\bnote\b", re.I)


def _infer_repeated_controls(steps: list, min_repeats: int = 2) -> int:
    """Treat a short name that recurs across the sheet as a control to click.

    In a services/POS walkthrough the same names come back page after page —
    "New Deal" x8, "Services Main Menu" x7, "Close Cash Drawer" x3. A phrase
    that short, that title-cased, repeating that often is a button or menu item
    being pressed, not prose.

    This is the one inference here that is a genuine guess, so it is deliberately
    narrow: Title Case, at most four words, no punctuation, and it must recur.
    A wrong click fails loudly in the report and is one dropdown away from being
    corrected in the editor, which beats asking the tester a question they
    cannot answer about a word with no context.
    """
    counts: dict = {}
    for st in steps:
        if (st.get("action") or "") != "todo":
            continue
        t = (st.get("target") or "").strip()
        # A section heading repeats too, but clicking "Buttons" is meaningless.
        if (2 <= len(t) <= 30 and _CONTROLish.match(t)
                and not _NOISE.search(t) and not _HEADING.match(t)):
            counts[t] = counts.get(t, 0) + 1
    repeated = {t for t, n in counts.items() if n >= min_repeats}
    changed = 0
    for st in steps:
        if (st.get("action") or "") != "todo":
            continue
        t = (st.get("target") or "").strip()
        if t in repeated:
            st["action"] = "click"
            st["target"] = t
            changed += 1
    return changed


def _render_yaml(doc: dict, comment_header: str) -> str:
    import yaml
    body = yaml.safe_dump(doc, sort_keys=False, default_flow_style=False,
                          allow_unicode=True, width=100)
    return comment_header + body


# ============================================================ Public API

@dataclass
class ImportResult:
    layout: str
    n_tests: int
    n_steps_total: int
    n_steps_translated: int           # everything that's NOT a `todo`
    cases: list                       # list[TestCase]   (legacy single-screen)
    yaml_text: str
    # New for multi-screen template imports:
    per_screen: dict | None = None    # {screen: yaml_text}
    screens: list | None = None       # ordered list of screens covered


def list_sheets(file_bytes: bytes) -> list[dict]:
    """Sheet names with a rough size, so the UI can offer a real choice."""
    import openpyxl, io
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    out = []
    for ws in wb.worksheets:
        rows = [[c.value for c in row] for row in ws.iter_rows()]
        while rows and not any(_norm(c) for c in rows[-1]):
            rows.pop()
        header = ""
        for r in rows[:8]:
            j = " ".join(_norm(c).lower() for c in r)
            if "step" in j or "scenario" in j or "feature" in j:
                header = " | ".join(_norm(c) for c in r if _norm(c))[:90]
                break
        out.append({"name": ws.title, "rows": max(0, len(rows) - 1),
                    "looks_like_tests": bool(header), "header": header})
    return out


def _pick_sheet(wb, sheet: str | None):
    """The named sheet, else the first that looks like test cases, else active."""
    if sheet:
        for ws in wb.worksheets:
            if ws.title == sheet:
                return ws
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
            j = " ".join(_norm(c).lower() for c in row)
            if ("step" in j and ("scenario" in j or "feature" in j or "test case" in j)):
                return ws
    return wb.active


def import_xlsx(file_bytes: bytes, screen: str = "yourscreen",
                sheet: str | None = None) -> ImportResult:
    """Parse an uploaded .xlsx into TestCases + render to YAML.

    Two paths:
      1. **Fast path** — official template (see framework/template_builder.py).
         Each row is one step. Multiple screens supported. 100% YAML.
      2. **Legacy auto-translation** — prose Excel files (the original
         Celerant test-case files). Uses regex rules; unmappable prose
         becomes `action: todo` markers.

    `screen` is the default screenname used by the legacy path and by any
    template row whose Screen cell is blank."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # Look for the template format across ALL worksheets — pick the one with
    # the most data rows. A workbook can have an empty "Test Cases" sheet
    # plus a filled "Example" sheet; we want the filled one.
    best_template = None    # (n_data_rows, rows, header_idx)
    for ws in wb.worksheets:
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        while rows and not any(_norm(c) for c in rows[-1]): rows.pop()
        if not rows: continue
        header_idx = _looks_like_template(rows)
        if header_idx is None: continue
        n_data = sum(1 for r in rows[header_idx + 1:] if any(_norm(c) for c in r))
        if n_data == 0: continue
        if best_template is None or n_data > best_template[0]:
            best_template = (n_data, rows, header_idx)

    if best_template is not None:
        _, rows, header_idx = best_template
        # ---------- FAST PATH ----------
        grouped = _parse_template(rows, header_idx)
        n_tests  = sum(len(sm) for sm in grouped.values())
        n_steps  = sum(len(tc.steps) for sm in grouped.values() for tc in sm.values())

        # Per-screen YAMLs
        per_screen: dict[str, str] = {}
        for scr, sm in sorted(grouped.items()):
            doc = {"tests": [
                {"screen": scr, "name": tc.name, "steps": [dict(st) for st in tc.steps]}
                for tc in sm.values()
            ]}
            header = (
                f"# ============================================================\n"
                f"#  Stratus QA — Imported from template (Excel)\n"
                f"#  Screen     : {scr}\n"
                f"#  Tests      : {len(sm)}\n"
                f"#  Steps      : {sum(len(tc.steps) for tc in sm.values())}\n"
                f"# ============================================================\n\n"
            )
            per_screen[scr] = _render_yaml(doc, header)

        # Combined YAML — single doc with EVERY screen's tests, useful
        # when the user wants one paste-into-editor blob.
        combined_doc = {"tests": []}
        for scr, sm in sorted(grouped.items()):
            for tc in sm.values():
                combined_doc["tests"].append(
                    {"screen": scr, "name": tc.name,
                     "steps": [dict(st) for st in tc.steps]})
        combined_header = (
            f"# ============================================================\n"
            f"#  Stratus QA — Imported from template (Excel)\n"
            f"#  Layout     : template (fast path)\n"
            f"#  Screens    : {len(grouped)}  ({', '.join(sorted(grouped))})\n"
            f"#  Test cases : {n_tests}\n"
            f"#  Steps      : {n_steps}\n"
            f"#  All steps map directly to actions — no prose translation.\n"
            f"# ============================================================\n\n"
        )
        combined_yaml = _render_yaml(combined_doc, combined_header)
        flat_cases = [tc for sm in grouped.values() for tc in sm.values()]
        return ImportResult(
            layout="template", n_tests=n_tests,
            n_steps_total=n_steps, n_steps_translated=n_steps,
            cases=flat_cases, yaml_text=combined_yaml,
            per_screen=per_screen, screens=sorted(grouped),
        )

    # ---------- LEGACY auto-translation ----------
    # wb.active is whichever tab Excel happened to have open when the file was
    # saved — for this workbook that was "Test Cases Services", so the sheet the
    # tester was actually looking at never got imported. Choose explicitly.
    ws = _pick_sheet(wb, sheet)
    rows = [[cell.value for cell in row] for row in ws.iter_rows()]
    while rows and not any(_norm(c) for c in rows[-1]): rows.pop()
    if not rows:
        return ImportResult("empty", 0, 0, 0, [], "tests: []\n")

    layout, header_idx = _detect_layout(rows)
    cols = _index_columns(rows[header_idx])
    if _looks_grouped(rows, cols, header_idx):
        layout = "grouped"; cases = _parse_grouped(rows, cols, header_idx)
    else:
        cases = _parse_flat(rows, cols, header_idx)

    # A screen's component list is documentation, but every element in it should
    # be on screen — so it converts to real assertions instead of questions.
    # These sheets put one row per line, so a heading and the elements under it
    # land in DIFFERENT cases (132 of 171 cases here hold a single step). The
    # grouping only survives in sheet order, so the pass runs across the whole
    # ordered sequence rather than inside each case.
    _flat = [st for c in cases for st in c.steps]
    _infer_inventory_checks(_flat)
    _infer_repeated_controls(_flat)

    doc = {"tests": [
        {"screen": screen,
         "name": c.name[:120],
         "steps": [dict(st) for st in c.steps] or [{"action": "todo", "target": "(no steps imported)"}]}
        for c in cases]}
    n_steps_total = sum(len(c.steps) for c in cases)
    n_translated  = sum(1 for c in cases for st in c.steps
                        if st.get("action") not in ("todo", "manual", "note"))
    pct = (100 * n_translated // n_steps_total) if n_steps_total else 0
    header = (
        f"# ============================================================\n"
        f"#  Stratus QA — Imported from Excel (legacy prose translation)\n"
        f"#  Screen          : {screen}\n"
        f"#  Layout detected : {layout}\n"
        f"#  Test cases      : {len(cases)}\n"
        f"#  Steps total     : {n_steps_total}\n"
        f"#  Auto-translated : {n_translated}  ({pct}%)\n"
        f"#  TODO markers    : {n_steps_total - n_translated}\n"
        f"# ------------------------------------------------------------\n"
        f"#  For 100% clean import use the official template:\n"
        f"#    docs/Stratus-QA-TestCase-Template.xlsx\n"
        f"# ============================================================\n\n"
    )
    return ImportResult(layout, len(cases), n_steps_total, n_translated,
                        cases, _render_yaml(doc, header))
